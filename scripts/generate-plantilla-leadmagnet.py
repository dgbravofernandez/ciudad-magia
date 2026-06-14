"""Genera una plantilla Excel profesional para inscripciones temporada que
sirve como lead magnet en la landing. El que la descarga deja email.
"""
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

out = Workbook()
out.remove(out.active)

# Estilos rosa Cluberly
header_fill = PatternFill(start_color='EC4899', end_color='EC4899', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
alt_fill = PatternFill(start_color='FDF2F8', end_color='FDF2F8', fill_type='solid')
border = Border(left=Side(style='thin', color='E5E7EB'), right=Side(style='thin', color='E5E7EB'),
                top=Side(style='thin', color='E5E7EB'), bottom=Side(style='thin', color='E5E7EB'))

# Hoja 1: Inscripciones
ws = out.create_sheet('Inscripciones')
headers = ['#', 'Nombre', 'Apellidos', 'Fecha nacimiento', 'DNI', 'Email tutor', 'Telefono tutor',
           'Categoria', 'Equipo asignado', 'Estado', 'Cuota base', 'Descuento hermano', 'Total a pagar',
           'Fecha pago reserva', 'Notas']
widths = [4, 18, 24, 16, 12, 28, 16, 16, 18, 14, 12, 16, 14, 16, 30]
for i, h in enumerate(headers, 1):
    c = ws.cell(row=1, column=i, value=h)
    c.fill = header_fill; c.font = header_font; c.border = border
    c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.column_dimensions[get_column_letter(i)].width = widths[i-1]
ws.row_dimensions[1].height = 32

# Ejemplos
examples = [
    (1, 'Lucía', 'Martín Pérez', '15/03/2015', '12345678A', 'tutor1@email.com', '600111222',
     'Alevín', 'Alevín A', 'Inscrito', 280, 0, 280, '01/07/2026', ''),
    (2, 'Mario', 'García López', '08/06/2014', '12345678B', 'tutor2@email.com', '600333444',
     'Infantil', 'Infantil B', 'Pendiente', 320, 0, 320, '', 'Falta foto DNI'),
    (3, 'Sofía', 'Ruiz Sánchez', '21/09/2016', '12345678C', 'tutor3@email.com', '600555666',
     'Benjamín', 'Benjamín C', 'Inscrito', 260, 26, 234, '03/07/2026', 'Hermana de #1, 10% descuento'),
]
for ri, row_vals in enumerate(examples, 2):
    for ci, v in enumerate(row_vals, 1):
        c = ws.cell(row=ri, column=ci, value=v)
        c.border = border
        if ri % 2 == 0: c.fill = alt_fill

# Filas vacías hasta 50
for ri in range(5, 51):
    for ci in range(1, len(headers) + 1):
        c = ws.cell(row=ri, column=ci)
        c.border = border
        if ri % 2 == 0: c.fill = alt_fill
    ws.cell(row=ri, column=1, value=ri-1)

ws.freeze_panes = 'B2'

# Hoja 2: Categorías y cuotas
ws2 = out.create_sheet('Categorias y cuotas')
cat_headers = ['Categoría', 'Año nacimiento', 'Cuota anual', 'Cuota mensual', 'Reserva']
for i, h in enumerate(cat_headers, 1):
    c = ws2.cell(row=1, column=i, value=h)
    c.fill = header_fill; c.font = header_font; c.border = border
    c.alignment = Alignment(horizontal='center')
    ws2.column_dimensions[get_column_letter(i)].width = 22
categories = [
    ('Prebenjamín', '2018-2019', 260, 26, 50),
    ('Benjamín', '2016-2017', 280, 28, 50),
    ('Alevín', '2014-2015', 320, 32, 50),
    ('Infantil', '2012-2013', 360, 36, 50),
    ('Cadete', '2010-2011', 400, 40, 50),
    ('Juvenil', '2008-2009', 420, 42, 50),
]
for ri, (cat, anyo, an, men, res) in enumerate(categories, 2):
    cells = [cat, anyo, f'{an} €', f'{men} €', f'{res} €']
    for ci, v in enumerate(cells, 1):
        c = ws2.cell(row=ri, column=ci, value=v)
        c.border = border
        if ri % 2 == 0: c.fill = alt_fill

# Hoja 3: Plan de cobros mensual
ws3 = out.create_sheet('Plan de cobros')
ws3['A1'] = 'PLAN DE COBROS — TEMPORADA 2026/2027'
ws3['A1'].font = Font(bold=True, size=14, color='BE185D')
ws3.merge_cells('A1:E1')
ws3['A3'] = 'Mes'
ws3['B3'] = 'Concepto'
ws3['C3'] = 'Cuotas previstas'
ws3['D3'] = 'Importe estimado'
ws3['E3'] = 'Acumulado'
for ci in range(1, 6):
    c = ws3.cell(row=3, column=ci)
    c.fill = header_fill; c.font = header_font; c.border = border
    ws3.column_dimensions[get_column_letter(ci)].width = 20

months = [
    ('Junio 26', 'Reserva (50€ por jugador)', 120, '6.000 €', '6.000 €'),
    ('Septiembre 26', 'Primera mensualidad', 120, '4.000 €', '10.000 €'),
    ('Octubre 26', 'Segunda mensualidad', 120, '4.000 €', '14.000 €'),
    ('Noviembre 26', 'Tercera mensualidad', 120, '4.000 €', '18.000 €'),
    ('Diciembre 26', 'Cuarta mensualidad', 120, '4.000 €', '22.000 €'),
    ('Enero 27', 'Quinta mensualidad', 120, '4.000 €', '26.000 €'),
    ('Febrero 27', 'Sexta mensualidad', 120, '4.000 €', '30.000 €'),
    ('Marzo 27', 'Séptima mensualidad', 120, '4.000 €', '34.000 €'),
    ('Abril 27', 'Octava mensualidad', 120, '4.000 €', '38.000 €'),
    ('Mayo 27', 'Última mensualidad', 120, '4.000 €', '42.000 €'),
]
for ri, row_vals in enumerate(months, 4):
    for ci, v in enumerate(row_vals, 1):
        c = ws3.cell(row=ri, column=ci, value=v)
        c.border = border
        if ri % 2 == 0: c.fill = alt_fill

# Hoja 4: INSTRUCCIONES + branding Cluberly
ws4 = out.create_sheet('Leer primero', 0)
ws4.column_dimensions['A'].width = 90
ws4['A1'] = 'PLANTILLA INSCRIPCIONES 2026/2027'
ws4['A1'].font = Font(bold=True, size=20, color='BE185D')

ws4['A3'] = 'Plantilla profesional preparada por Cluberly — software de gestión para clubes amateur.'
ws4['A3'].font = Font(size=12, italic=True, color='6B7280')

texts = [
    (5, 'COMO USAR ESTA PLANTILLA', Font(bold=True, size=14, color='1F2937')),
    (7, '1. Pestaña "Inscripciones" — rellena cada jugador en una fila.', None),
    (8, '2. Pestaña "Categorias y cuotas" — ajusta los importes de tu club.', None),
    (9, '3. Pestaña "Plan de cobros" — visualiza el plan mensual de tesorería.', None),
    (11, 'PROBLEMAS QUE NO RESUELVE UN EXCEL', Font(bold=True, size=14, color='1F2937')),
    (13, '✗ Familias que pagan tarde: tienes que perseguirlas a mano por WhatsApp.', None),
    (14, '✗ Recordatorios automáticos: imposible. Te encargas tú una por una.', None),
    (15, '✗ Avisos de deuda con justificante: tienes que redactarlos cada vez.', None),
    (16, '✗ Cuadre con banco: Excel no concilia transferencias automáticamente.', None),
    (17, '✗ Cuando lo abre varias personas: descuadres, conflictos, duplicados.', None),
    (19, 'Si esto te suena: prueba Cluberly 14 días gratis, sin tarjeta.', Font(bold=True, color='EC4899', size=12)),
    (20, 'https://cluberly.club', Font(color='BE185D', underline='single', size=12)),
    (22, 'Plan Básico — 39€/mes hasta 100 jugadores', None),
    (23, 'Plan Pro — 89€/mes hasta 300 jugadores', None),
    (24, 'Plan Club — 149€/mes hasta 600 jugadores', None),
    (26, 'O reserva 15 min de demo: https://cluberly.club/reservar', Font(color='BE185D', underline='single')),
    (28, 'Diego Bravo — Cluberly', Font(italic=True, color='6B7280', size=10)),
    (29, 'iakevoapp@gmail.com · 665 676 341', Font(italic=True, color='6B7280', size=10)),
]
for cell_row, text, font in texts:
    ws4.cell(row=cell_row, column=1, value=text)
    if font:
        ws4[f'A{cell_row}'].font = font

# Guardar
os.makedirs('public/downloads', exist_ok=True)
output_path = 'public/downloads/plantilla-inscripciones-2026-2027.xlsx'
out.save(output_path)
print(f'OK: {output_path} ({os.path.getsize(output_path) // 1024} KB)')
