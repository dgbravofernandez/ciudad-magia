"""Genera Excel multi-federación con:
- Pestaña INSTRUCCIONES con plan operativo
- Pestaña RFFM Madrid con los 628 clubes ya conseguidos
- Pestaña por cada federación autonómica restante (vacías, con cabecera + cómo conseguir datos)
"""
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

# 1. Leer el Excel original RFFM
src = openpyxl.load_workbook('docs/marketing/Seguimiento-Clubes-Cluberly.xlsx', data_only=True)
ws_src = src['Seguimiento']
rffm_data = list(ws_src.iter_rows(values_only=True))
print(f"RFFM source: {len(rffm_data)} filas (incluida cabecera)")

# 2. Federaciones target
FEDERATIONS = [
    {'sheet': 'RFFM Madrid',         'web': 'https://rffm.com',                 'estim': '750',  'access': 'Intranet RFFM (acceso clubes federados)',           'tier': '1'},
    {'sheet': 'FFM Futbol Sala',     'web': 'https://ffmadrid.es',              'estim': '300',  'access': 'Contactar info@ffmadrid.es',                        'tier': '1'},
    {'sheet': 'FCF Cataluna',        'web': 'https://www.fcf.cat',              'estim': '1500', 'access': 'Contactar partnerships@fcf.cat',                    'tier': '1'},
    {'sheet': 'RFAF Andalucia',      'web': 'https://rfaf.es',                  'estim': '1500', 'access': 'Contactar info@rfaf.es',                            'tier': '1'},
    {'sheet': 'FFCV Valencia',       'web': 'https://ffcv.es',                  'estim': '1000', 'access': 'Contactar federacion@ffcv.es',                      'tier': '1'},
    {'sheet': 'FFRM Murcia',         'web': 'https://ffrm.es',                  'estim': '250',  'access': 'Contactar info@ffrm.es',                            'tier': '2'},
    {'sheet': 'FGF Galicia',         'web': 'https://futgal.es',                'estim': '500',  'access': 'Contactar info@futgal.es',                          'tier': '2'},
    {'sheet': 'FFPA Asturias',       'web': 'https://fapas.es',                 'estim': '250',  'access': 'Contactar info@fapas.es',                           'tier': '2'},
    {'sheet': 'FFCM Castilla LM',    'web': 'https://fferm.es',                 'estim': '500',  'access': 'Contactar info@fferm.es',                           'tier': '2'},
    {'sheet': 'FCYLF Castilla Leon', 'web': 'https://fcylf.es',                 'estim': '600',  'access': 'Contactar info@fcylf.es',                           'tier': '2'},
    {'sheet': 'FAF Aragon',          'web': 'https://www.aragonfutbol.com',     'estim': '300',  'access': 'Contactar info@aragonfutbol.com',                   'tier': '2'},
    {'sheet': 'FFIB Baleares',       'web': 'https://www.ffib.es',              'estim': '200',  'access': 'Contactar info@ffib.es',                            'tier': '2'},
    {'sheet': 'IFCF Canarias',       'web': 'https://www.fifcanaria.es',        'estim': '300',  'access': 'Contactar info@fifcanaria.es',                      'tier': '2'},
    {'sheet': 'FFV Vasca',           'web': 'https://www.eff-fvf.eus',          'estim': '400',  'access': 'Contactar info@eff-fvf.eus',                        'tier': '2'},
    {'sheet': 'FNF Navarra',         'web': 'https://www.ffnavarra.com',        'estim': '150',  'access': 'Contactar info@ffnavarra.com',                      'tier': '3'},
    {'sheet': 'FCF Cantabria',       'web': 'https://www.fcf-cantabria.com',    'estim': '150',  'access': 'Contactar info@fcf-cantabria.com',                  'tier': '3'},
    {'sheet': 'FRF La Rioja',        'web': 'https://www.fedrioja.com',         'estim': '100',  'access': 'Contactar info@fedrioja.com',                       'tier': '3'},
    {'sheet': 'FEXF Extremadura',    'web': 'https://www.fexfutbol.com',        'estim': '250',  'access': 'Contactar info@fexfutbol.com',                      'tier': '3'},
]

# 3. Crear nuevo workbook
out = Workbook()
out.remove(out.active)

# Estilos
header_fill = PatternFill(start_color='EC4899', end_color='EC4899', fill_type='solid')
header_font = Font(bold=True, color='FFFFFF', size=11)
border = Border(left=Side(style='thin', color='E5E7EB'), right=Side(style='thin', color='E5E7EB'),
                top=Side(style='thin', color='E5E7EB'), bottom=Side(style='thin', color='E5E7EB'))
alt_fill = PatternFill(start_color='FDF2F8', end_color='FDF2F8', fill_type='solid')

# 4. Hoja INSTRUCCIONES
ws = out.create_sheet('INSTRUCCIONES')
ws.column_dimensions['A'].width = 22
ws.column_dimensions['B'].width = 14
ws.column_dimensions['C'].width = 38
ws.column_dimensions['D'].width = 55
ws.column_dimensions['E'].width = 10

ws['A1'] = 'CLUBERLY - Plan de captacion multi-federacion'
ws['A1'].font = Font(bold=True, size=14, color='BE185D')
ws.merge_cells('A1:E1')

instructions = [
    ('A3', 'COMO USAR ESTE EXCEL', Font(bold=True, size=12)),
    ('A4', '1) RFFM Madrid ya esta completo con 628 clubes (los que tenias en el Excel original).', None),
    ('A5', '2) Para cada federacion nueva, hay una pestana lista con la cabecera correcta.', None),
    ('A6', '3) Para conseguir los datos de cada federacion (legalmente):', None),
    ('A7', '   a) Contacta por email a la federacion pidiendo partnership o convenio.', None),
    ('A8', '   b) Mensaje sugerido: "Soy Diego de Cluberly, software de gestion para clubes amateur.', None),
    ('A9', '      Estoy abriendo colaboracion con federaciones autonomicas. Posible reunion?"', None),
    ('A10', '   c) Si te dan Excel, subelo en cluberly.club/superadmin/leads-import.', None),
    ('A11', '   d) Alternativa: Google Maps Places API por provincia para encontrar clubes amateur.', None),
    ('A13', 'Total potencial estimado: ~8.250 clubes amateur en Espana.', Font(bold=True, color='BE185D')),
    ('A14', 'Si capturas el 5% = 412 clientes potenciales. Con 30% close rate = 124 clientes pagando.', Font(color='10B981')),
]
for cell, value, font in instructions:
    ws[cell] = value
    if font:
        ws[cell].font = font

# Tabla federaciones
row = 17
headers = ['Federacion', 'Estimacion', 'Web oficial', 'Como conseguir datos', 'Tier']
for i, h in enumerate(headers, 1):
    c = ws.cell(row=row, column=i, value=h)
    c.fill = header_fill
    c.font = header_font
    c.border = border
    c.alignment = Alignment(horizontal='center', vertical='center')

tier_colors = {'1': '10B981', '2': 'F59E0B', '3': '6B7280'}
for i, fed in enumerate(FEDERATIONS):
    row += 1
    cells_data = [fed['sheet'], fed['estim'], fed['web'], fed['access'], fed['tier']]
    for j, val in enumerate(cells_data, 1):
        c = ws.cell(row=row, column=j, value=val)
        c.border = border
        if i % 2:
            c.fill = alt_fill
        if j == 1:
            c.font = Font(bold=True)
        if j == 5:
            c.font = Font(bold=True, color=tier_colors.get(val, '000000'))
            c.alignment = Alignment(horizontal='center')

# 5. Una hoja por federacion con cabecera del original
RFFM_HEADERS = list(rffm_data[0])
print(f"Cabeceras del original: {RFFM_HEADERS}")

for fed in FEDERATIONS:
    ws = out.create_sheet(fed['sheet'])
    # Cabecera
    for i, h in enumerate(RFFM_HEADERS, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill = header_fill
        c.font = header_font
        c.border = border
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[1].height = 28

    for i in range(1, len(RFFM_HEADERS) + 1):
        ws.column_dimensions[get_column_letter(i)].width = 22
    ws.column_dimensions['A'].width = 32
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['N'].width = 35

    # RFFM: pegar datos reales
    if fed['sheet'] == 'RFFM Madrid':
        for row_idx, row_data in enumerate(rffm_data[1:], 2):
            for col_idx, val in enumerate(row_data, 1):
                c = ws.cell(row=row_idx, column=col_idx, value=val)
                c.border = border
                if row_idx % 2 == 0:
                    c.fill = alt_fill
        # Banner final
        total_row = len(rffm_data) + 2
        c = ws.cell(row=total_row, column=1, value=f'TOTAL: {len(rffm_data) - 1} clubes RFFM (ya importados a Cluberly)')
        c.font = Font(bold=True, color='10B981')
    else:
        # Banner explicativo arriba
        info_row = 3
        c = ws.cell(row=info_row, column=1, value=f'PENDIENTE - Estimacion: {fed["estim"]} clubes')
        c.font = Font(bold=True, color='EC4899', size=12)

        c = ws.cell(row=info_row + 1, column=1, value=f'Conseguir: {fed["access"]}')
        c.font = Font(color='6B7280')

        c = ws.cell(row=info_row + 2, column=1, value=f'Web: {fed["web"]}')
        c.font = Font(color='6B7280')

        # Fila ejemplo
        example_row = info_row + 4
        example_values = [
            'Ej: C.D. Ejemplo',                  # Club
            'Localidad',                          # Ubicacion
            f'{fed["web"]}/clubejemplo',         # Web
            'club@ejemplo.es',                    # Email
            '6XX XXX XXX',                        # Telefono
            fed['sheet'].split()[0],              # Federacion
            'Nuevo',                              # Estado
            '',                                   # Persona contacto
            '',                                   # Fecha 1er contacto
            '0',                                  # Emails enviados
            '',                                   # Fecha ultimo contacto
            'Enviar Email 1',                     # Proxima accion
            '',                                   # Fecha prox. accion
            f'Pendiente conseguir desde {fed["access"]}',  # Notas
        ]
        for col_idx, val in enumerate(example_values, 1):
            c = ws.cell(row=example_row, column=col_idx, value=val)
            c.border = border
            c.font = Font(italic=True, color='9CA3AF')
            c.fill = PatternFill(start_color='F9FAFB', end_color='F9FAFB', fill_type='solid')

# Guardar
output_path = 'docs/marketing/Seguimiento-Clubes-Multi-Federacion.xlsx'
out.save(output_path)
sz = os.path.getsize(output_path)
print(f"\nGUARDADO: {output_path} ({sz // 1024} KB)")
print(f"Hojas creadas: INSTRUCCIONES + {len(FEDERATIONS)} federaciones")
print(f"RFFM con datos: {len(rffm_data) - 1} clubes")
print(f"Resto: pestanas estructuradas listas para rellenar")
