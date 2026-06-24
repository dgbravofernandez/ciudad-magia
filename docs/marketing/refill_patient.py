"""
refill_patient.py  v2.0 — PNFG scraper mejorado
Mejoras vs v1:
  - 20+ User-Agents modernos (Chrome/Firefox/Edge/Safari 2024-2025)
  - Headers completos de navegador real (Sec-Fetch-*, Sec-CH-UA, etc.)
  - Session warm-up: visita homepage antes del listing (cookies + anti-bot)
  - CB_SLEEP adaptativo: crece exponencialmente (10→15→20→25 min) no fijo
  - Jitter aleatorio ±30% en todos los sleeps para evitar patrones detectables
  - Retry en errores de red/timeout, no solo en respuesta vacía
  - parse_email mejorado: busca mailto: links + texto completo además de h5
  - Rotate UA cada 20 fichas (antes 30)
  - Skip automático de federaciones ya 100% cubiertas
  - Blacklist de emails ampliada
"""
import os, re, sys, time, random, unicodedata, socket
from urllib.parse import urlparse, urljoin

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

import openpyxl, requests
from bs4 import BeautifulSoup
import urllib3
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

sys.path.insert(0, os.path.dirname(__file__))
from scraper_federaciones import (
    FEDERACIONES, COL_CLUB, COL_EMAIL, COL_TEL,
    atomic_save,
)

COL_WEB   = 3
RUTA      = r"C:\Users\dgbra\Ciudad Magia\docs\marketing\Seguimiento-Clubes-Multi-Federacion.xlsx"

# ─── TIEMPOS BASE (se aplica jitter ±30% sobre todos) ────────────────────────
# v3.0 — ROUND-ROBIN + BAIL RÁPIDO
#   Causa raíz del atasco en ~50 emails/fed: los servidores PNFG pequeños
#   (novanet compartido) bloquean tras ~50-80 requests rápidos. El CB v2 hacía
#   6 trips con sleeps de hasta 45 min (2h muertas) y luego el early-exit
#   confundía "bloqueado" con "sin emails" → abortaba feds que SÍ tienen emails.
#   v3: tras 2 backoffs cortos, ABANDONA esta fed por este round (no la pierde),
#   salta a la siguiente (otro dominio, que está fresco) y vuelve después de que
#   el primero se haya enfriado. El early-exit ahora mira SOLO fichas que SÍ
#   cargaron (loaded), nunca las bloqueadas.
SLEEP_FICHA    = 5.0   # pace moderado por ficha
SLEEP_PAGE     = 6.0
SLEEP_FED      = 8.0    # pausa entre federaciones (cambio de dominio)
SAVE_EVERY     = 10

# Round-robin: máximo de fichas por fed en cada pasada. Al alcanzarlo, salta a la
# siguiente fed para que este dominio se enfríe mientras trabajamos otros.
MAX_PER_CALL   = 80

# Detección de fed genuinamente SIN emails (no confundir con bloqueo):
# solo si cargaron >=50 fichas OK y el ratio de email es < 5%.
MIN_LOADED_FOR_NOEMAIL = 50
NOEMAIL_RATE           = 0.05

def jitter(base: float, pct: float = 0.25) -> float:
    """Aplica ±pct de variación aleatoria para evitar patrones detectables."""
    return base * random.uniform(1 - pct, 1 + pct)

# CIRCUIT BREAKER — backoffs CORTOS, bail rápido (no quemar horas)
CB_THRESHOLD   = 8     # 8 bloqueos consecutivos = rate-limited ahora mismo
CB_MAX_TRIPS   = 2     # tras 2 backoffs, abandona esta fed este round (resume luego)

def cb_sleep_time(trips: int) -> float:
    """Backoff corto: 90s → 180s. Si sigue bloqueando, mejor saltar de fed."""
    bases = [90, 180]
    return float(bases[min(trips - 1, len(bases) - 1)])

ROTATE_EVERY   = 20   # rotar UA cada 20 fichas

# Overrides por federación (sobre la config base de scraper_federaciones.py)
FED_VERIFY   = {"FEXF Extremadura": False}   # cert SSL roto → verify=False
FED_DISABLED = {                              # plataforma rota: necesitan browser-automation
    "FFIB Baleares",      # homepage len=121, PNFG movido de host
    "FVF Euskadi",        # instancia PNFG distinta, no lista clubes por cod
    # NOTA: Castilla LM (ffcm.es) NO está deshabilitada: el probe dio len=0 por
    # bloqueo temporal de IP (probes repetidos), pero el scraper en vivo saca
    # emails con normalidad (~29% hit sobre 1795 clubes). Funciona.
}

# Federaciones objetivo (baja cobertura)
TARGET_FEDS = {
    "FVF Euskadi",
    "RFAF Andalucia",
    "FFCM Castilla LM",
    "FEXF Extremadura",
    "IFCF Canarias",
    "FAF Aragon",
    "FFRM Murcia",
    "FFIB Baleares",
    "FNF Navarra",
    "FFPA Asturias",
    "FCYLF Castilla Leon",
    "FCF Cantabria",
    "FGF Galicia",
    "FRF La Rioja",
    "RFMF Melilla",
}

# ─── USER AGENTS modernos 2024-2025 ──────────────────────────────────────────
USER_AGENTS = [
    # Chrome Windows
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 11.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36",
    # Chrome Mac
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 14_5) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36",
    # Firefox Windows
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:127.0) Gecko/20100101 Firefox/127.0",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:126.0) Gecko/20100101 Firefox/126.0",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:125.0) Gecko/20100101 Firefox/125.0",
    # Firefox Mac/Linux
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 14.5; rv:127.0) Gecko/20100101 Firefox/127.0",
    "Mozilla/5.0 (X11; Ubuntu; Linux x86_64; rv:126.0) Gecko/20100101 Firefox/126.0",
    # Edge Windows
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36 Edg/126.0.0.0",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36 Edg/125.0.0.0",
    # Safari Mac
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 14_5) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.5 Safari/605.1.15",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 14_4_1) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.4.1 Safari/605.1.15",
    # Safari iPhone
    "Mozilla/5.0 (iPhone; CPU iPhone OS 17_5 like Mac OS X) AppleWebKit/605.1.15 (KHTML, like Gecko) Version/17.5 Mobile/15E148 Safari/604.1",
    # Opera
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/125.0.0.0 Safari/537.36 OPR/111.0.0.0",
    # Chrome Android
    "Mozilla/5.0 (Linux; Android 14; Pixel 8) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Mobile Safari/537.36",
    # Brave
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0.0.0 Safari/537.36 Brave/126",
    # Older but common
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
]

EMAIL_RE = re.compile(r"[\w.\-+]+@[\w.\-]+\.[a-z]{2,}", re.I)
EMAIL_BLACKLIST = {
    "contacto@ffcm.es", "info@rfaf.es", "info@rfef.es",
    "info@ffrm.es", "info@futgal.es", "contacto@rfcylf.es",
    "info@asturfutbol.es", "info@futnavarra.es", "info@ffib.es",
    "info@federacioncanariafutbol.es", "contacto@fexfutbol.es",
    "info@futbolaragon.com", "contact@example.com", "noreply@example.com",
    "club@ejemplo.es",
}


class AbortFederation(Exception):
    pass


def norm(s):
    if not s: return ""
    s = unicodedata.normalize('NFKD', s)
    s = ''.join(c for c in s if not unicodedata.combining(c))
    return re.sub(r'\s+', ' ', re.sub(r'[.,;:"\'\(\)\[\]/\\]+', ' ', s.lower())).strip()


def norm_words(s):
    STOP = {'cd','sd','ad','ud','ed','cf','sc','ac','fc','fs','rcd','rcf',
            'club','sociedad','asociacion','deportivo','deportiva','futbol',
            'sala','de','del','la','las','el','los','union','agrupacion','ce','ec',
            'at','atletico'}
    return {w for w in norm(s).split() if w and w not in STOP and len(w) >= 3}


def build_headers(ua: str, referer: str = "") -> dict:
    """Construye headers completos que imitan un navegador real."""
    is_firefox = "Firefox" in ua
    is_safari = "Safari" in ua and "Chrome" not in ua

    h = {
        "User-Agent": ua,
        "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,image/avif,image/webp,*/*;q=0.8",
        "Accept-Language": "es-ES,es;q=0.9,en;q=0.8",
        "Accept-Encoding": "gzip, deflate, br",
        "Connection": "keep-alive",
        "Upgrade-Insecure-Requests": "1",
        "Cache-Control": "max-age=0",
    }
    if referer:
        h["Referer"] = referer

    if not is_firefox and not is_safari:
        # Headers Chromium
        h["Sec-Fetch-Dest"] = "document"
        h["Sec-Fetch-Mode"] = "navigate"
        h["Sec-Fetch-Site"] = "same-origin" if referer else "none"
        h["Sec-Fetch-User"] = "?1"
        h["Sec-CH-UA"] = '"Chromium";v="126", "Not/A)Brand";v="8"'
        h["Sec-CH-UA-Mobile"] = "?0"
        h["Sec-CH-UA-Platform"] = '"Windows"'
    return h


def make_fresh_session(base_url: str, ua_idx: int = 0, verify: bool = True) -> requests.Session:
    """
    Crea sesión con warm-up real: visita homepage primero, luego la URL base.
    Esto inicializa cookies anti-bot correctamente.
    """
    ua = USER_AGENTS[ua_idx % len(USER_AGENTS)]
    s = requests.Session()
    s.headers.update(build_headers(ua))
    s.verify = verify   # algunas feds tienen cert SSL roto (Extremadura)

    # Warm-up 1: visita homepage del dominio
    parsed = urlparse(base_url)
    homepage = f"{parsed.scheme}://{parsed.netloc}/"
    try:
        s.get(homepage, timeout=20)
        time.sleep(jitter(1.5))
    except Exception:
        pass

    # Warm-up 2: visita la URL base (listing principal)
    try:
        s.headers.update(build_headers(ua, referer=homepage))
        s.get(base_url, timeout=20)
        time.sleep(jitter(1.0))
    except Exception:
        pass

    return s


def get_soup_with_retry(url: str, session: requests.Session,
                        retries: int = 3, timeout: int = 25,
                        verify: bool = True) -> BeautifulSoup | None:
    """
    GET con retry en errores de red/timeout (no solo en respuesta vacía).
    """
    last_err = None
    for attempt in range(retries):
        try:
            if attempt > 0:
                time.sleep(jitter(5.0 * attempt))  # backoff entre reintentos
            r = session.get(url, timeout=timeout, verify=verify)
            r.raise_for_status()
            r.encoding = r.apparent_encoding
            if len(r.text) >= 500:
                return BeautifulSoup(r.text, "html.parser")
            # Respuesta demasiado corta = posible bloqueo
            if attempt < retries - 1:
                time.sleep(jitter(10.0))
        except (requests.exceptions.Timeout,
                requests.exceptions.ConnectionError) as e:
            last_err = e
            # Retry en errores de red
        except requests.exceptions.HTTPError as e:
            last_err = e
            if hasattr(e, 'response') and e.response is not None:
                if e.response.status_code in (429, 503):
                    # Rate limit o servicio no disponible — esperar más
                    time.sleep(jitter(30.0 * (attempt + 1)))
            break  # otros errores HTTP no tienen retry
        except Exception as e:
            last_err = e
            break
    return None


def parse_email_enhanced(soup: BeautifulSoup) -> str:
    """
    Búsqueda de email mejorada:
      1. h5 > strong con label "email" (patrón PNFG estándar)
      2. Links mailto:
      3. Regex en todo el texto visible de la página
    """
    if not soup:
        return ""

    # 1. Patrón PNFG estándar: <h5><strong>Email:</strong> xxx@xxx.es</h5>
    for h5 in soup.find_all("h5"):
        strong = h5.find("strong")
        if strong and "email" in strong.get_text(strip=True).lower():
            text = h5.get_text(separator=" ", strip=True)
            m = EMAIL_RE.search(text)
            if m:
                return m.group(0).strip().lower()

    # 2. Links mailto: en la página
    for a in soup.find_all("a", href=re.compile(r"^mailto:", re.I)):
        href = a.get("href", "")
        m = EMAIL_RE.search(href)
        if m:
            return m.group(0).strip().lower()

    # 3. Regex en todo el texto visible
    page_text = soup.get_text(separator=" ")
    candidates = EMAIL_RE.findall(page_text)
    for email in candidates:
        email = email.strip().lower()
        # Filtrar emails de sistema/librerías
        if any(bad in email for bad in ['wixpress', 'wordpress', 'example',
                                         'sentry', 'jquery', '.png', '.jpg',
                                         '.gif', '.css', '.js']):
            continue
        if len(email) > 6 and '.' in email.split('@')[-1]:
            return email

    return ""


def parse_phone_enhanced(soup: BeautifulSoup) -> str:
    """Parse teléfono con patrón estándar PNFG."""
    if not soup:
        return ""
    for h5 in soup.find_all("h5"):
        strong = h5.find("strong")
        if not strong:
            continue
        label = strong.get_text(strip=True).lower()
        if re.search(r"tel|tfno|telf|movil|m[oó]vil|phone|fax", label):
            text = h5.get_text(separator=" ", strip=True)
            m = re.search(r"(\+?[\d][\d\s\-\.]{6,})", text)
            if m:
                return re.sub(r"\s+", " ", m.group(1)).strip()
    return ""


def build_mapping_with_retry(fed: dict, session: requests.Session, verify: bool = True):
    """Build club→ref mapping, retry hasta 4 veces con pausa si está vacío."""
    for attempt in range(4):
        if attempt > 0:
            print(f"  [map retry #{attempt}] Pausa {120 * attempt}s antes de reintentar listing...")
            time.sleep(jitter(120.0 * attempt))
            session = make_fresh_session(
                fed["list_url"] + f"?cod_primaria={fed['cod_primaria']}",
                attempt, verify=verify
            )

        pairs = []
        page = 1
        while True:
            url = (f"{fed['list_url']}?cod_primaria={fed['cod_primaria']}"
                   f"&NPcd_PageLines=999&NPcd_Page={page}&Buscar=1")
            soup = get_soup_with_retry(url, session, verify=verify)
            if not soup:
                break
            found = 0
            if fed["tipo"] == "A":
                for a in soup.find_all("a", href=re.compile(r"javascript:Ver\(\d+\)", re.I)):
                    nombre = a.get_text(strip=True)
                    m = re.search(r"Ver\((\d+)\)", a["href"], re.I)
                    if nombre and m:
                        pairs.append((nombre, m.group(1)))
                        found += 1
            else:
                for a in soup.find_all("a", href=re.compile(r"NFG_VisEquipos", re.I)):
                    nombre = a.get_text(strip=True)
                    href = a.get("href", "")
                    if nombre and href:
                        if not href.startswith("http"):
                            href = fed.get("base_domain", "") + href
                        pairs.append((nombre, href))
                        found += 1
            print(f"  [map] pagina {page}: {found} clubes")
            if found == 0:
                break
            page += 1
            time.sleep(jitter(SLEEP_PAGE))

        if pairs:
            return pairs, session
    return [], session


class FuzzyMatcher:
    def __init__(self, pairs):
        self.exact = {}
        self.full = []
        for nombre, ref in pairs:
            n = norm(nombre)
            self.exact.setdefault(n, ref)
            self.full.append((n, norm_words(nombre), ref))

    def find(self, query):
        q = norm(query)
        if not q: return (None, 'none')
        if q in self.exact: return (self.exact[q], 'exact')
        if len(q) >= 8:
            for n, _, ref in self.full:
                if q in n or n in q:
                    return (ref, 'contains')
        qw = norm_words(query)
        if len(qw) >= 2:
            best_ref, best_score = None, 0
            for _, nw, ref in self.full:
                if not nw: continue
                inter = len(qw & nw)
                if inter >= 2:
                    score = inter / len(qw | nw)
                    if score > best_score and score >= 0.5:
                        best_score = score
                        best_ref = ref
            if best_ref: return (best_ref, 'overlap')
        return (None, 'none')


# Cachés a nivel de módulo (persisten entre rounds del mismo proceso):
_MATCHER_CACHE: dict = {}   # hoja -> FuzzyMatcher | None (None = listing roto)
_ATTEMPTED: dict = {}       # hoja -> set(row_idx) ya intentados este proceso (evita re-fetch)


def get_matcher(fed: dict, verify: bool):
    """Construye (y cachea) el mapping club→ref de una federación."""
    hoja = fed['hoja']
    if hoja in _MATCHER_CACHE:
        return _MATCHER_CACHE[hoja]
    base_url = fed["list_url"] + f"?cod_primaria={fed['cod_primaria']}"
    print(f"  Construyendo mapping desde {fed['list_url']}...")
    session = make_fresh_session(base_url, 0, verify=verify)
    pairs, _ = build_mapping_with_retry(fed, session, verify=verify)
    print(f"  Mapping: {len(pairs)} entradas")
    matcher = FuzzyMatcher(pairs) if pairs else None
    _MATCHER_CACHE[hoja] = matcher
    return matcher


def refill_fed(fed: dict, wb, ws, max_n: int = MAX_PER_CALL) -> dict:
    """
    Procesa hasta `max_n` fichas sin email de una federación (round-robin friendly).
    Devuelve {'encontrados': N, 'status': ...} donde status ∈:
      'done'     → no quedan filas PNFG por rellenar (fed completa para PNFG)
      'more'     → alcanzó el cap; quedan filas para el próximo round
      'blocked'  → servidor bloqueó; abandona este round, resume luego
      'no_email' → las fichas cargan OK pero la fed no publica emails
    """
    hoja = fed['hoja']
    verify = FED_VERIFY.get(hoja, True)
    print(f"\n{'='*70}\n>> {hoja}  (tipo {fed['tipo']})\n{'='*70}")

    # Filas sin email (resume natural: lee el estado actual del Excel)
    attempted = _ATTEMPTED.setdefault(hoja, set())
    rows_sin: list[tuple[int, str, str]] = []
    for r in range(2, ws.max_row + 1):
        nombre = ws.cell(row=r, column=COL_CLUB).value
        email  = ws.cell(row=r, column=COL_EMAIL).value
        web    = ws.cell(row=r, column=COL_WEB).value
        if not nombre: continue
        n = str(nombre).strip()
        if not n or n in ('-', 'None'): continue
        e = str(email).strip() if email else ""
        if e in ("", "-", "None", "club@ejemplo.es"):
            w = str(web).strip() if web else ""
            if w in ("", "-", "None"): w = ""
            rows_sin.append((r, n, w))

    if not rows_sin:
        print("  Sin filas que rellenar — federacion completa.")
        return {'encontrados': 0, 'status': 'done'}

    total_filas = ws.max_row - 1
    cob = (total_filas - len(rows_sin)) / total_filas * 100 if total_filas else 0
    print(f"  SIN email: {len(rows_sin)} / {total_filas}  (cobertura {cob:.0f}%)  "
          f"— hasta {max_n} fichas este round")

    matcher = get_matcher(fed, verify)
    if matcher is None:
        print("  [SKIP] Listing vacio — servidor bloqueado o config rota.")
        return {'encontrados': 0, 'status': 'blocked'}

    base_url = fed["list_url"] + f"?cod_primaria={fed['cod_primaria']}"
    session = make_fresh_session(base_url, 0, verify=verify)

    total = since_save = 0
    loaded = hits = 0
    consecutive_blocks = trips = ua = done = 0
    status = 'done'   # si recorremos todo rows_sin sin cortar antes

    for row_idx, nombre, website in rows_sin:
        if done >= max_n:
            status = 'more'
            break
        if row_idx in attempted:
            continue                      # ya intentado este proceso (evita re-fetch)
        ref, modo = matcher.find(nombre)
        if not ref:
            continue                      # sin match en el listado → lo deja para web-enrich

        url = (f"{fed['ficha_url']}?cod_primaria={fed['cod_primaria']}&codigo_club={ref}"
               if fed["tipo"] == "A" else ref)
        time.sleep(jitter(SLEEP_FICHA))
        soup = get_soup_with_retry(url, session, verify=verify)
        done += 1

        if soup is None:
            # BLOQUEO (no "sin email"): la web no devolvió ficha válida
            consecutive_blocks += 1
            if consecutive_blocks >= CB_THRESHOLD:
                trips += 1
                if trips > CB_MAX_TRIPS:
                    print(f"  [CB] {trips} backoffs sin recuperar — abandono {hoja} "
                          f"este round (resume luego).")
                    status = 'blocked'
                    break
                s = cb_sleep_time(trips)
                print(f"  [CB #{trips}] {CB_THRESHOLD} bloqueos seguidos. "
                      f"Backoff {s:.0f}s + sesion nueva...")
                time.sleep(s)
                ua += 1
                session = make_fresh_session(base_url, ua, verify=verify)
                consecutive_blocks = 0
            continue

        # Ficha cargó OK
        consecutive_blocks = 0
        loaded += 1
        attempted.add(row_idx)
        if done % ROTATE_EVERY == 0:
            ua += 1
            session = make_fresh_session(base_url, ua, verify=verify)

        email = parse_email_enhanced(soup)
        phone = parse_phone_enhanced(soup)
        if email and email.lower() in EMAIL_BLACKLIST:
            email = ""
        updated = False
        if email:
            ws.cell(row=row_idx, column=COL_EMAIL).value = email
            hits += 1; updated = True
        if phone:
            ws.cell(row=row_idx, column=COL_TEL).value = phone
            updated = True
        if updated:
            total += 1; since_save += 1
        if updated or done % 10 == 0:
            rate = hits / loaded * 100 if loaded else 0
            print(f"  [{done}/{min(max_n, len(rows_sin))}] [{modo}] {nombre[:32]:<32} "
                  f"email: {email or '-':<26} tel: {phone or '-'}  (hit {rate:.0f}%)")

        # Fed genuinamente SIN emails: fichas cargan OK pero casi ninguna trae email.
        # (Esto NO se confunde con bloqueo: solo cuenta fichas que cargaron.)
        if loaded >= MIN_LOADED_FOR_NOEMAIL and (hits / loaded) < NOEMAIL_RATE:
            print(f"  [NO-EMAIL] {loaded} fichas OK, solo {hits} con email "
                  f"({hits/loaded*100:.0f}%) — esta fed no publica emails en ficha.")
            status = 'no_email'
            break

        if since_save >= SAVE_EVERY:
            atomic_save(wb, RUTA); since_save = 0

    atomic_save(wb, RUTA)
    rate = hits / loaded * 100 if loaded else 0
    print(f"  -- {hoja}: +{total} emails este round  "
          f"(fichas OK {loaded}, hit {rate:.0f}%, status={status})")
    return {'encontrados': total, 'status': status}


# ─── DB CHECK ─────────────────────────────────────────────────────────────────

def check_email_count_in_db() -> int:
    try:
        url = "https://mcjmguvkcseyfhsyvdhd.supabase.co/rest/v1/marketing_clubs"
        key = ("eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI"
               "6Im1jam1ndXZrY3NleWZoc3l2ZGhkIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI"
               "6MTc3NTA3MjczNCwiZXhwIjoyMDkwNjQ4NzM0fQ.1nBVmV2tQQyOj9EzER1UEXHEv2C"
               "-oIND_ItZIrNpcQs")
        h = {"apikey": key, "Authorization": f"Bearer {key}", "Prefer": "count=exact"}
        r = requests.get(url + "?email=not.is.null&select=id",
                         headers=h, params={"limit": 1}, timeout=10)
        return int(r.headers.get("content-range", "0/0").split("/")[-1])
    except Exception as e:
        print(f"  [DB] Error: {e}")
        return 0


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def run_import_now(tag: str = "") -> None:
    """Importa Excel → Supabase (idempotente: on_conflict=email). Datos a salvo aunque maten el proceso."""
    import subprocess
    print(f"\n  [sync] Importando a Supabase ({tag})...")
    try:
        r = subprocess.run(
            [sys.executable, "import_to_supabase.py"],
            capture_output=True, text=True, encoding='utf-8',
            cwd=os.path.dirname(os.path.abspath(__file__)),
            timeout=180,
        )
        if r.stdout:
            for line in r.stdout.strip().splitlines()[-4:]:
                print(f"    {line}")
        if r.returncode != 0 and r.stderr:
            print(f"    [WARN import] {r.stderr[-200:]}")
    except Exception as e:
        print(f"    [WARN import] {e}")


def sheet_coverage(ws) -> float:
    """% de filas con email en una hoja."""
    total_filas = ws.max_row - 1
    if total_filas <= 0:
        return 100.0
    sin = sum(
        1 for r in range(2, ws.max_row + 1)
        if str(ws.cell(row=r, column=COL_EMAIL).value or "").strip()
        in ("", "-", "None", "club@ejemplo.es")
    )
    return (total_filas - sin) / total_filas * 100


def main():
    # Prioridad por nº de clubs sin email (mayor primero)
    priority = {
        "RFAF Andalucia": 2640, "IFCF Canarias": 853, "FEXF Extremadura": 973,
        "FAF Aragon": 539, "FFRM Murcia": 421, "FFPA Asturias": 260,
        "FNF Navarra": 274, "FCYLF Castilla Leon": 217, "FCF Cantabria": 116,
        "FGF Galicia": 41, "RFMF Melilla": 41, "FRF La Rioja": 5,
    }
    # Excluir feds deshabilitadas (plataforma rota) y hojas inexistentes
    socket.setdefaulttimeout(45)
    print(f"Abriendo Excel: {RUTA}")
    wb = openpyxl.load_workbook(RUTA)

    feds_to_run = [
        f for f in FEDERACIONES
        if f["hoja"] in TARGET_FEDS
        and f["hoja"] not in FED_DISABLED
        and f["hoja"] in wb.sheetnames
    ]
    feds_to_run.sort(key=lambda f: priority.get(f["hoja"], 0), reverse=True)
    if FED_DISABLED:
        print(f"[INFO] Deshabilitadas (necesitan browser-automation): "
              f"{', '.join(sorted(FED_DISABLED))}")

    # Presupuesto de tiempo por proceso (override con SCRAPER_BUDGET_H). Round-robin:
    # cada round procesa hasta MAX_PER_CALL fichas/fed y salta a la siguiente, dando
    # tiempo a cada dominio a enfriarse antes de volver.
    start = time.time()
    TIME_BUDGET = float(os.environ.get("SCRAPER_BUDGET_H", "3.5")) * 3600
    MAX_ROUNDS  = 15
    completed: set = set()
    no_email:  set = set()
    grand_total = 0

    for round_n in range(1, MAX_ROUNDS + 1):
        active = [f for f in feds_to_run
                  if f["hoja"] not in completed and f["hoja"] not in no_email]
        if not active:
            print("\n[FIN] No quedan federaciones activas.")
            break
        print(f"\n{'#'*70}\n# ROUND {round_n}/{MAX_ROUNDS} — {len(active)} feds activas\n{'#'*70}")
        gained = 0
        time_up = False
        for fed in active:
            if time.time() - start > TIME_BUDGET:
                print(f"\n[TIME] Presupuesto {TIME_BUDGET/3600:.1f}h agotado — paro.")
                time_up = True
                break
            ws = wb[fed["hoja"]]
            cob = sheet_coverage(ws)
            if cob >= 95.0:
                print(f"\n[SKIP] {fed['hoja']} — cobertura {cob:.0f}% >= 95%.")
                completed.add(fed["hoja"])
                continue
            res = refill_fed(fed, wb, ws)
            grand_total += res['encontrados']
            gained += res['encontrados']
            if res['encontrados'] > 0:
                run_import_now(fed["hoja"])
            st = res['status']
            if st == 'done':
                completed.add(fed["hoja"])
            elif st == 'no_email':
                no_email.add(fed["hoja"])
            # 'more' / 'blocked' → sigue activa para el próximo round (resume)
            time.sleep(jitter(SLEEP_FED))
        if time_up:
            break
        if gained == 0 and round_n >= 2:
            print("\n[FIN] Round sin progreso — restantes bloqueadas o sin match PNFG.")
            break

    wb.close()

    print(f"\n{'='*70}")
    print(f"TOTAL emails nuevos este proceso: {grand_total}")
    if completed:
        print(f"Completadas: {', '.join(sorted(completed))}")
    if no_email:
        print(f"Sin emails en ficha (web-enrich aparte): {', '.join(sorted(no_email))}")
    print(f"{'='*70}")

    run_import_now("FINAL")
    count = check_email_count_in_db()
    print(f"\n>>> EMAILS CON EMAIL EN BD AHORA: {count:,}")


if __name__ == "__main__":
    main()
