"""
import_to_supabase.py v2
Lee todas las hojas de Seguimiento-Clubes-Multi-Federacion.xlsx
e importa TODOS los clubes a marketing_clubs en Supabase.

- Con email:    upsert con on_conflict=email (ignora duplicados)
- Sin email:    INSERT solo si (name_lower, fed_lower) no existe ya en BD,
                status='no_email' (no entran al pipeline de envio)
"""
import re
import sys
import time
import requests

# ─── CONFIG ───────────────────────────────────────────────────────────────────
EXCEL_PATH   = r"C:\Users\dgbra\Ciudad Magia\docs\marketing\Seguimiento-Clubes-Multi-Federacion.xlsx"
SUPABASE_URL = "https://mcjmguvkcseyfhsyvdhd.supabase.co"
SERVICE_KEY  = "eyJhbGciOiJIUzI1NiIsInR5cCI6IkpXVCJ9.eyJpc3MiOiJzdXBhYmFzZSIsInJlZiI6Im1jam1ndXZrY3NleWZoc3l2ZGhkIiwicm9sZSI6InNlcnZpY2Vfcm9sZSIsImlhdCI6MTc3NTA3MjczNCwiZXhwIjoyMDkwNjQ4NzM0fQ.1nBVmV2tQQyOj9EzER1UEXHEv2C-oIND_ItZIrNpcQs"
CHUNK_SIZE   = 200
# Hojas placeholder sin datos reales — saltar
SKIP_SHEETS  = {"INSTRUCCIONES", "FFM Futbol Sala", "FCF Cataluna", "FFCV Valencia", "FFV Vasca"}

# ─── HELPERS ──────────────────────────────────────────────────────────────────
EMAIL_RE = re.compile(r'^[^\s@]+@[^\s@]+\.[^\s@]+$')

HEADERS = {
    "apikey":        SERVICE_KEY,
    "Authorization": f"Bearer {SERVICE_KEY}",
    "Content-Type":  "application/json",
}

def valid_email(s):
    return bool(s and EMAIL_RE.match(s.strip().lower()))

def auto_excluded(name):
    n = name.upper()
    return bool(
        re.search(r'\bS\.?A\.?D\.?\b', n) or
        re.search(r'\bS\.?L\.?\b', n) or
        'REAL MADRID' in n or
        ('ATLETICO' in n and 'MADRID' in n) or
        'BARCELONA' in n
    )

def auto_priority(name):
    n = name.upper()
    if re.search(r'\bESCUELA\b', n) or re.search(r'\bA\.D\b', n) or \
       re.search(r'\bC\.D\b', n) or re.search(r'\bE\.F\b', n):
        return 10
    return 100

def upsert_email_batch(rows):
    """Upsert clubs con email, ignora duplicados."""
    url = f"{SUPABASE_URL}/rest/v1/marketing_clubs?on_conflict=email"
    r = requests.post(url, json=rows,
                      headers={**HEADERS, "Prefer": "resolution=ignore-duplicates,return=representation"},
                      timeout=30)
    if r.status_code not in (200, 201):
        print(f"    [ERROR upsert] {r.status_code}: {r.text[:200]}")
        return 0
    return len(r.json()) if r.text.strip().startswith('[') else 0

def insert_no_email_batch(rows):
    """INSERT clubs sin email. Si el batch falla (409 duplicado), reintenta uno a uno."""
    if not rows:
        return 0
    url = f"{SUPABASE_URL}/rest/v1/marketing_clubs"
    r = requests.post(url, json=rows,
                      headers={**HEADERS, "Prefer": "return=minimal"},
                      timeout=30)
    if r.status_code in (200, 201):
        return len(rows)
    if r.status_code == 409:
        # Batch contiene duplicado(s) — reintentar fila a fila
        inserted = 0
        for row in rows:
            r2 = requests.post(url, json=[row],
                               headers={**HEADERS, "Prefer": "return=minimal"},
                               timeout=15)
            if r2.status_code in (200, 201):
                inserted += 1
            # 409 = duplicado → silencioso (ya existe en BD)
        return inserted
    print(f"    [ERROR insert_no_email] {r.status_code}: {r.text[:200]}")
    return 0

_NO_EMAIL_CACHE: set | None = None

def load_existing_no_email(_imported_from=None):
    """Carga (una vez, paginado) todos los clubes sin email de BD → set de (name_lower, fed_lower)."""
    global _NO_EMAIL_CACHE
    if _NO_EMAIL_CACHE is not None:
        return _NO_EMAIL_CACHE
    _NO_EMAIL_CACHE = set()
    offset = 0
    page_size = 1000
    while True:
        url = (f"{SUPABASE_URL}/rest/v1/marketing_clubs"
               f"?select=name,federation&email=is.null"
               f"&limit={page_size}&offset={offset}")
        r = requests.get(url, headers=HEADERS, timeout=30)
        if r.status_code != 200:
            print(f"    [WARN] Error cargando sin email (offset={offset}): {r.status_code}")
            break
        batch = r.json()
        if not batch:
            break
        for row in batch:
            n = (row.get('name') or '').lower().strip()
            f = (row.get('federation') or '').lower().strip()
            _NO_EMAIL_CACHE.add((n, f))
        if len(batch) < page_size:
            break
        offset += page_size
    print(f"  [dedup cache] {len(_NO_EMAIL_CACHE)} clubes sin email cargados de BD")
    return _NO_EMAIL_CACHE

# ─── MAIN ──────────────────────────────────────────────────────────────────────
try:
    import openpyxl
except ImportError:
    print("pip install openpyxl")
    sys.exit(1)

print(f"Abriendo Excel: {EXCEL_PATH}")
wb = openpyxl.load_workbook(EXCEL_PATH, read_only=True, data_only=True)

total_email_new = 0
total_no_email_new = 0
total_skipped = 0

for sheet_name in wb.sheetnames:
    if sheet_name in SKIP_SHEETS:
        continue

    ws = wb[sheet_name]
    rows_iter = ws.iter_rows(values_only=True)
    header_row = next(rows_iter, None)
    if header_row is None:
        continue

    # Columnas: A=Club, B=Ubicacion, C=Web, D=Email, E=Telefono, F=Federacion
    col_name  = 0
    col_loc   = 1
    col_web   = 2
    col_email = 3
    col_phone = 4
    col_fed   = 5

    # Cargar existentes sin email para dedup manual
    existing_no_email = load_existing_no_email(sheet_name)

    batch_email    = []
    batch_no_email = []
    sheet_email_new    = 0
    sheet_no_email_new = 0
    sheet_skip     = 0

    for row in rows_iter:
        name = str(row[col_name]).strip() if row[col_name] else ''
        if not name or name in ('-', 'None'):
            sheet_skip += 1
            continue

        raw_email = str(row[col_email]).strip().lower() if row[col_email] else ''
        email_ok  = valid_email(raw_email) and raw_email != 'club@ejemplo.es'

        loc   = str(row[col_loc]).strip()   if row[col_loc]   and str(row[col_loc])   not in ('-', 'None') else None
        web   = str(row[col_web]).strip()   if row[col_web]   and str(row[col_web])   not in ('-', 'None') else None
        phone = str(row[col_phone]).strip() if row[col_phone] and str(row[col_phone]) not in ('-', 'None') else None
        fed_raw = str(row[col_fed]).strip() if row[col_fed] and str(row[col_fed]) not in ('-', 'None') else sheet_name

        excl  = auto_excluded(name)
        prio  = auto_priority(name)

        base = {
            "name":          name[:200],
            "location":      loc[:100]  if loc  else None,
            "federation":    (fed_raw[:100] if fed_raw else sheet_name),
            "website":       web[:300]  if web  else None,
            "phone":         phone[:100] if phone else None,
            "priority":      prio,
            "excluded":      excl,
            "imported_from": sheet_name,
            "notes":         "Auto-excluido: profesional/SAD" if excl else None,
        }

        if email_ok:
            batch_email.append({**base, "email": raw_email[:200]})
            if len(batch_email) >= CHUNK_SIZE:
                n = upsert_email_batch(batch_email)
                sheet_email_new += n
                batch_email = []
        else:
            # Dedup manual: saltamos si ya existe en BD
            key = (name.lower().strip(), (fed_raw or sheet_name).lower().strip())
            if key in existing_no_email:
                sheet_skip += 1
                continue
            existing_no_email.add(key)  # evitar duplicados dentro del mismo import
            batch_no_email.append({**base, "email": None, "status": "no_email"})
            if len(batch_no_email) >= CHUNK_SIZE:
                n = insert_no_email_batch(batch_no_email)
                sheet_no_email_new += n
                batch_no_email = []

    # Flush remanentes
    if batch_email:
        n = upsert_email_batch(batch_email)
        sheet_email_new += n
    if batch_no_email:
        n = insert_no_email_batch(batch_no_email)
        sheet_no_email_new += n

    total_email_new    += sheet_email_new
    total_no_email_new += sheet_no_email_new
    total_skipped      += sheet_skip
    print(f"  {sheet_name:<35} +email={sheet_email_new:>4}  +sin_email={sheet_no_email_new:>5}  (skip={sheet_skip})")

wb.close()

print()
print(f"TOTAL CON EMAIL (nuevos/upsert): {total_email_new}")
print(f"TOTAL SIN EMAIL (nuevos):        {total_no_email_new}")
print(f"TOTAL SALTADOS:                  {total_skipped}")
print(f"GRAN TOTAL PROCESADOS:           {total_email_new + total_no_email_new}")
