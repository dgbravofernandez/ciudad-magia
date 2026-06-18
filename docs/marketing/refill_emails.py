"""
refill_emails.py
Recorre el Excel buscando clubes SIN email y re-consulta la ficha del PNFG
para rellenar email + telefono.

3 mejoras vs version inicial:
  1. Normalizacion + fuzzy match del nombre (sin tildes, sin puntos, sin abreviaturas)
  2. Cola de reintentos al final de cada federacion con pausa de 5 min (rate-limit temporal)
  3. Fallback: si tras todo no hay email pero hay website, scrapear el website del club

Uso:
    python refill_emails.py                       # todas las federaciones PNFG
    python refill_emails.py "RFAF Andalucia"      # solo una
    python refill_emails.py "Andalucia" "Murcia"  # varias
"""
import os
import re
import sys
import time
import unicodedata
from urllib.parse import urlparse

if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

import openpyxl
import requests
from bs4 import BeautifulSoup

# Importar config del scraper original
sys.path.insert(0, os.path.dirname(__file__))
from scraper_federaciones import (
    FEDERACIONES, HEADERS, COL_CLUB, COL_EMAIL, COL_TEL,
    make_session, parse_email, parse_phone, atomic_save,
)

# Columna C en el Excel = website (1-indexed = 3)
COL_WEB = 3

RUTA = r"C:\Users\dgbra\Ciudad Magia\docs\marketing\Seguimiento-Clubes-Multi-Federacion.xlsx"

# Tiempos
SLEEP_FICHA     = 3.0
SLEEP_PAGE      = 3.0
SLEEP_RETRY     = 20.0
SLEEP_FED       = 10.0
SLEEP_RETRY_Q   = 300.0
MAX_RETRIES     = 0       # sin retry inline - el circuit breaker se encarga
SAVE_EVERY      = 30

# CIRCUIT BREAKER: si K vacios seguidos -> pausa larga + rotar session
CB_THRESHOLD    = 5       # 5 fichas vacias consecutivas activa el breaker
CB_SLEEP        = 180.0   # 3 min de pausa cuando dispara
CB_MAX_TRIPS    = 3       # tras 3 trips en una federacion, esa fed va a la cola final
ROTATE_EVERY    = 50      # renovar session + cambiar UA cada 50 fichas

USER_AGENTS = [
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/121.0.0.0 Safari/537.36",
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64; rv:120.0) Gecko/20100101 Firefox/120.0",
    "Mozilla/5.0 (X11; Linux x86_64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/122.0.0.0 Safari/537.36",
]

EMAIL_RE = re.compile(r"[\w.\-+]+@[\w.\-]+\.[a-z]{2,}", re.I)


class AbortFederation(Exception):
    """Señal: el servidor de esta federacion esta bloqueando. Saltar a la siguiente."""

# Emails que NO son contactos reales del club (boilerplate del PNFG)
EMAIL_BLACKLIST = {
    "contacto@ffcm.es",
    "info@rfaf.es",
    "info@rfef.es",
}


# ─── NORMALIZACION ───────────────────────────────────────────────────────────

def norm(s: str) -> str:
    """Normaliza nombre de club para fuzzy matching:
       - sin tildes
       - lowercase
       - sin puntos / comas / comillas
       - sin abreviaturas "C.D.", "S.D.", etc -> "cd", "sd"
       - colapsa espacios
    """
    if not s:
        return ""
    s = unicodedata.normalize('NFKD', s)
    s = ''.join(c for c in s if not unicodedata.combining(c))
    s = s.lower()
    s = re.sub(r'[.,;:"\'\(\)\[\]/\\]+', ' ', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s


def norm_words(s: str) -> set[str]:
    """Devuelve set de palabras significativas (sin stopwords) para fuzzy."""
    STOP = {'cd','sd','ad','ud','ed','cf','sc','ac','fc','fs','rcd','rcf',
            'club','sociedad','asociacion','deportivo','deportiva','futbol',
            'sala','de','del','la','las','el','los','union','agrupacion',
            'ce','ec','at','atletico'}
    return {w for w in norm(s).split() if w and w not in STOP and len(w) >= 3}


# ─── HTTP ────────────────────────────────────────────────────────────────────

def get_soup_patient(url: str, session) -> BeautifulSoup | None:
    """Sin retry inline - el circuit breaker en el caller maneja los vacios."""
    try:
        r = session.get(url, timeout=20)
        r.raise_for_status()
        r.encoding = r.apparent_encoding
        if len(r.text) >= 500:
            return BeautifulSoup(r.text, "html.parser")
    except Exception:
        pass
    return None


def make_session_random(base_url: str, ua_idx: int = 0) -> requests.Session:
    """Sesion nueva con UA distinto - parte del IP rotation light."""
    s = requests.Session()
    headers = dict(HEADERS)
    headers["User-Agent"] = USER_AGENTS[ua_idx % len(USER_AGENTS)]
    s.headers.update(headers)
    try:
        s.get(base_url, timeout=15)
    except Exception:
        pass
    return s


# ─── MAPPING NOMBRE -> CODIGO (con fuzzy match) ──────────────────────────────

def build_mapping_tipo_a(fed, session) -> list[tuple[str, str]]:
    """Devuelve lista [(nombre_original, codigo)] del listado del PNFG."""
    out: list[tuple[str, str]] = []
    page = 1
    while True:
        url = (f"{fed['list_url']}?cod_primaria={fed['cod_primaria']}"
               f"&NPcd_PageLines=999&NPcd_Page={page}&Buscar=1")
        soup = get_soup_patient(url, session)
        if not soup:
            break
        found = 0
        for a in soup.find_all("a", href=re.compile(r"javascript:Ver\(\d+\)", re.I)):
            nombre = a.get_text(strip=True)
            m = re.search(r"Ver\((\d+)\)", a["href"], re.I)
            if nombre and m:
                out.append((nombre, m.group(1)))
                found += 1
        print(f"  [map] pagina {page}: {found} clubes")
        if found == 0:
            break
        page += 1
        time.sleep(SLEEP_PAGE)
    return out


def build_mapping_tipo_b(fed, session) -> list[tuple[str, str]]:
    out: list[tuple[str, str]] = []
    page = 1
    while True:
        url = (f"{fed['list_url']}?cod_primaria={fed['cod_primaria']}"
               f"&NPcd_PageLines=999&NPcd_Page={page}&Buscar=1")
        soup = get_soup_patient(url, session)
        if not soup:
            break
        found = 0
        for a in soup.find_all("a", href=re.compile(r"NFG_VisEquipos", re.I)):
            nombre = a.get_text(strip=True)
            href = a.get("href", "")
            if not nombre or not href:
                continue
            if not href.startswith("http"):
                href = fed["base_domain"] + href
            out.append((nombre, href))
            found += 1
        print(f"  [map] pagina {page}: {found} clubes")
        if found == 0:
            break
        page += 1
        time.sleep(SLEEP_PAGE)
    return out


class FuzzyMatcher:
    """Match exact -> contiene -> overlap de palabras significativas."""
    def __init__(self, pairs: list[tuple[str, str]]):
        # exact: norm(nombre) -> ref
        self.exact: dict[str, str] = {}
        # full: lista de (norm(nombre), set_palabras, ref) para fuzzy
        self.full: list[tuple[str, set[str], str]] = []
        for nombre, ref in pairs:
            n = norm(nombre)
            self.exact.setdefault(n, ref)
            self.full.append((n, norm_words(nombre), ref))

    def find(self, query: str) -> tuple[str | None, str]:
        """Devuelve (ref, modo_match) - modo: 'exact'/'contains'/'overlap'/'none'."""
        q = norm(query)
        if not q:
            return (None, 'none')
        # 1. Exact match
        if q in self.exact:
            return (self.exact[q], 'exact')
        # 2. Substring (un lado contiene al otro)
        if len(q) >= 8:
            for n, _, ref in self.full:
                if q in n or n in q:
                    return (ref, 'contains')
        # 3. Overlap por palabras significativas
        qw = norm_words(query)
        if len(qw) >= 2:
            best_ref = None
            best_score = 0
            for _, nw, ref in self.full:
                if not nw:
                    continue
                # Jaccard de palabras significativas
                inter = len(qw & nw)
                if inter >= 2:
                    union = len(qw | nw)
                    score = inter / union
                    if score > best_score and score >= 0.5:
                        best_score = score
                        best_ref = ref
            if best_ref:
                return (best_ref, 'overlap')
        return (None, 'none')


# ─── FALLBACK: scrapear el website propio del club ───────────────────────────

def scrape_email_from_website(url: str) -> str:
    """Intenta extraer un email del propio website del club.
       Visita home + /contacto + /contact. Devuelve el PRIMER email no-genérico encontrado.
    """
    if not url:
        return ""
    if not url.startswith(('http://', 'https://')):
        url = 'https://' + url
    try:
        parsed = urlparse(url)
        if not parsed.netloc:
            return ""
        base = f"{parsed.scheme}://{parsed.netloc}"
    except Exception:
        return ""

    candidates = [
        url,
        base,
        base + '/contacto',
        base + '/contact',
        base + '/contacto/',
        base + '/contacto.html',
    ]
    found_emails = set()
    for cand in candidates[:4]:  # max 4 peticiones por club
        try:
            r = requests.get(cand, headers=HEADERS, timeout=10, allow_redirects=True)
            if r.status_code != 200:
                continue
            r.encoding = r.apparent_encoding
            text = r.text
            for m in EMAIL_RE.finditer(text):
                e = m.group(0).lower()
                # Filtrar emails de proveedores/genericos del HTML (no del club)
                if any(bad in e for bad in [
                    'wixpress', 'wordpress.com', 'sentry', 'github',
                    'example.com', 'dominio.com', 'tudominio',
                    'jpg', 'png', 'css', 'svg',
                ]):
                    continue
                if e in EMAIL_BLACKLIST:
                    continue
                found_emails.add(e)
            if found_emails:
                break  # ya tenemos suficiente
        except Exception:
            continue
        time.sleep(0.5)

    if not found_emails:
        return ""
    # Preferir emails del propio dominio del club
    domain = parsed.netloc.lower().replace('www.', '')
    for e in found_emails:
        if domain.split('.')[0] in e.split('@')[1]:
            return e
    # Si no, el primero que no sea claramente boilerplate
    return sorted(found_emails)[0]


# ─── PROCESO PRINCIPAL POR FEDERACION ────────────────────────────────────────

def refill_federation(fed, wb, ws) -> dict:
    """Devuelve dict con stats."""
    print(f"\n{'='*70}")
    print(f">> {fed['hoja']}  (tipo {fed['tipo']})")
    print(f"{'='*70}")

    # Encontrar filas SIN email
    rows_sin_email: list[tuple[int, str, str]] = []  # (row_idx, nombre, website)
    for row_idx in range(2, ws.max_row + 1):
        nombre = ws.cell(row=row_idx, column=COL_CLUB).value
        email = ws.cell(row=row_idx, column=COL_EMAIL).value
        web = ws.cell(row=row_idx, column=COL_WEB).value
        if not nombre:
            continue
        nombre_str = str(nombre).strip()
        if not nombre_str or nombre_str in ('-', 'None', 'club@ejemplo.es'):
            continue
        email_str = str(email).strip() if email else ""
        if email_str in ("", "-", "None"):
            web_str = str(web).strip() if web else ""
            if web_str in ("", "-", "None"):
                web_str = ""
            rows_sin_email.append((row_idx, nombre_str, web_str))

    if not rows_sin_email:
        print(f"  Sin filas que rellenar. Saltando.")
        return {'procesados': 0, 'encontrados': 0, 'sin_resultado': 0}

    print(f"  Filas SIN email: {len(rows_sin_email)}")

    # Construir mapping del PNFG
    base = fed["list_url"] + f"?cod_primaria={fed['cod_primaria']}"
    session = make_session(base)
    print(f"  Construyendo mapping desde {fed['list_url']}...")
    if fed["tipo"] == "A":
        pairs = build_mapping_tipo_a(fed, session)
    else:
        pairs = build_mapping_tipo_b(fed, session)
    print(f"  Mapping construido: {len(pairs)} entradas")
    if not pairs:
        print(f"  [WARN] Mapping vacio - servidor bloquea. Skip federacion.")
        return {'procesados': 0, 'encontrados': 0, 'sin_resultado': 0}

    matcher = FuzzyMatcher(pairs)

    stats = {
        'procesados': 0,
        'encontrados_pnfg_exact': 0,
        'encontrados_pnfg_fuzzy': 0,
        'encontrados_web': 0,
        'sin_resultado': 0,
        'no_match_nombre': 0,
    }
    encontrados_total = 0
    since_save = 0
    retry_queue: list[tuple[int, str, str, str]] = []
    # Circuit breaker state (scoped to this federation)
    cb_state = {'empties_in_row': 0, 'trips': 0, 'session': session, 'ua_idx': 0, 'fichas_done': 0}

    def attempt_ficha(row_idx, nombre, website, ref, modo: str) -> bool:
        """Intenta llenar la ficha. Devuelve True si encontro email/tel.
           Aplica circuit breaker + rotacion de session/UA."""
        nonlocal encontrados_total, since_save
        if fed["tipo"] == "A":
            ficha_url = (f"{fed['ficha_url']}?cod_primaria={fed['cod_primaria']}"
                         f"&codigo_club={ref}")
        else:
            ficha_url = ref
        time.sleep(SLEEP_FICHA)
        soup_ficha = get_soup_patient(ficha_url, cb_state['session'])
        cb_state['fichas_done'] += 1

        # Circuit breaker: si demasiados vacios seguidos -> pausar y rotar
        if soup_ficha is None:
            cb_state['empties_in_row'] += 1
            if cb_state['empties_in_row'] >= CB_THRESHOLD:
                cb_state['trips'] += 1
                if cb_state['trips'] > CB_MAX_TRIPS:
                    print(f"  [CB] Demasiados trips ({cb_state['trips']}). Abortando federacion - va a cola final.")
                    raise AbortFederation()
                print(f"  [CB trip #{cb_state['trips']}] {CB_THRESHOLD} vacios seguidos. Pausa {CB_SLEEP/60:.0f} min + rotando UA...")
                time.sleep(CB_SLEEP)
                cb_state['ua_idx'] += 1
                cb_state['session'] = make_session_random(base, cb_state['ua_idx'])
                cb_state['empties_in_row'] = 0
            return False
        else:
            # Reset on success
            cb_state['empties_in_row'] = 0

        # Rotacion periodica de session/UA
        if cb_state['fichas_done'] % ROTATE_EVERY == 0:
            cb_state['ua_idx'] += 1
            cb_state['session'] = make_session_random(base, cb_state['ua_idx'])
        email = parse_email(soup_ficha)
        phone = parse_phone(soup_ficha)
        # filtrar boilerplate de la propia federacion
        if email and email.lower() in EMAIL_BLACKLIST:
            email = ""
        updated = False
        if email:
            ws.cell(row=row_idx, column=COL_EMAIL).value = email
            updated = True
            if modo == 'exact':
                stats['encontrados_pnfg_exact'] += 1
            else:
                stats['encontrados_pnfg_fuzzy'] += 1
        if phone:
            ws.cell(row=row_idx, column=COL_TEL).value = phone
            updated = True
        if updated:
            encontrados_total += 1
            since_save += 1
            print(f"  [{stats['procesados']}/{len(rows_sin_email)}] [{modo}] {nombre[:38]:<38} "
                  f"email: {email or '-':<28} tel: {phone or '-'}")
        return bool(email)

    # ─── PASE 1: fuzzy match contra PNFG ─────────────────────────────────────
    aborted_pase1 = False
    i = -1
    try:
        for i, (row_idx, nombre, website) in enumerate(rows_sin_email):
            stats['procesados'] += 1
            ref, modo = matcher.find(nombre)
            if not ref:
                stats['no_match_nombre'] += 1
                retry_queue.append((row_idx, nombre, website, ''))
                continue
            ok = attempt_ficha(row_idx, nombre, website, ref, modo)
            if not ok:
                retry_queue.append((row_idx, nombre, website, ref))
            if since_save >= SAVE_EVERY:
                atomic_save(wb, RUTA)
                print(f"  [save] Excel guardado ({encontrados_total} encontrados)")
                since_save = 0
    except AbortFederation:
        aborted_pase1 = True
        # Meter el resto sin procesar en retry_queue
        for j in range(i+1, len(rows_sin_email)):
            row_idx, nombre, website = rows_sin_email[j]
            stats['procesados'] += 1
            ref, modo = matcher.find(nombre)
            retry_queue.append((row_idx, nombre, website, ref if ref else ''))
            if not ref:
                stats['no_match_nombre'] += 1

    atomic_save(wb, RUTA)
    print(f"  -- Pase 1 {'(ABORTADO por CB)' if aborted_pase1 else 'OK'}: {encontrados_total} encontrados, {len(retry_queue)} pendientes")

    # ─── PASE 2: cola de reintentos PNFG con pausa larga ─────────────────────
    # Solo los que tenian ref (estaban en el listing pero ficha dio vacia)
    pnfg_retries = [(ri, n, w, r) for ri, n, w, r in retry_queue if r]
    web_only     = [(ri, n, w) for ri, n, w, r in retry_queue if not r]

    if pnfg_retries:
        print(f"\n  [PASE 2] Pausa {SLEEP_RETRY_Q/60:.0f} min antes de reintentar {len(pnfg_retries)} fichas vacias...")
        time.sleep(SLEEP_RETRY_Q)
        # Renovar session completamente con UA distinto
        cb_state['ua_idx'] += 1
        cb_state['session'] = make_session_random(base, cb_state['ua_idx'])
        cb_state['empties_in_row'] = 0
        cb_state['trips'] = 0
        try:
            for row_idx, nombre, website, ref in pnfg_retries:
                ok = attempt_ficha(row_idx, nombre, website, ref, 'retry')
                if not ok:
                    web_only.append((row_idx, nombre, website))
                if since_save >= SAVE_EVERY:
                    atomic_save(wb, RUTA)
                    since_save = 0
        except AbortFederation:
            print(f"  [CB] Pase 2 abortado tambien. Los pendientes van directos a web fallback.")
        atomic_save(wb, RUTA)
        print(f"  -- Pase 2 OK: total acumulado {encontrados_total}, {len(web_only)} pendientes para web fallback")

    # ─── PASE 3: fallback al website propio del club ─────────────────────────
    web_candidates = [(ri, n, w) for ri, n, w in web_only if w]
    if web_candidates:
        print(f"\n  [PASE 3] Web fallback en {len(web_candidates)} clubes con sitio propio...")
        for row_idx, nombre, website in web_candidates:
            try:
                email = scrape_email_from_website(website)
            except Exception:
                email = ""
            if email:
                ws.cell(row=row_idx, column=COL_EMAIL).value = email
                stats['encontrados_web'] += 1
                encontrados_total += 1
                since_save += 1
                print(f"  [WEB] {nombre[:38]:<38} email: {email}")
                if since_save >= SAVE_EVERY:
                    atomic_save(wb, RUTA)
                    since_save = 0
            time.sleep(1)
        atomic_save(wb, RUTA)

    stats['sin_resultado'] = stats['procesados'] - encontrados_total
    print(f"\n  -- RESUMEN {fed['hoja']}:")
    print(f"     PNFG exact: {stats['encontrados_pnfg_exact']}, fuzzy: {stats['encontrados_pnfg_fuzzy']}, "
          f"web: {stats['encontrados_web']}, sin_match: {stats['no_match_nombre']}, "
          f"total: {encontrados_total}/{stats['procesados']}")
    return stats


def count_missing_emails(ws) -> int:
    n = 0
    for row_idx in range(2, ws.max_row + 1):
        nombre = ws.cell(row=row_idx, column=COL_CLUB).value
        email = ws.cell(row=row_idx, column=COL_EMAIL).value
        if not nombre:
            continue
        email_str = str(email).strip() if email else ""
        if email_str in ("", "-", "None"):
            n += 1
    return n


def main():
    filtro = [a.lower() for a in sys.argv[1:]]
    print(f"Abriendo {RUTA}")
    wb = openpyxl.load_workbook(RUTA)

    feds_orden = []
    for fed in FEDERACIONES:
        if filtro and not any(f in fed["hoja"].lower() for f in filtro):
            continue
        if fed["hoja"] not in wb.sheetnames:
            print(f"[WARN] Hoja '{fed['hoja']}' no en el Excel - skip")
            continue
        ws = wb[fed["hoja"]]
        missing = count_missing_emails(ws)
        feds_orden.append((missing, fed))

    feds_orden.sort(key=lambda x: -x[0])
    print("\nPLAN (huecos por federacion, mayor primero):")
    total_huecos = 0
    for missing, fed in feds_orden:
        print(f"  {fed['hoja']:30} {missing:5} huecos")
        total_huecos += missing
    print(f"  {'TOTAL':30} {total_huecos:5} huecos\n")

    resumen = []
    for missing, fed in feds_orden:
        if missing == 0:
            continue
        ws = wb[fed["hoja"]]
        try:
            stats = refill_federation(fed, wb, ws)
            resumen.append((fed["hoja"], stats))
        except Exception as exc:
            print(f"  [FATAL] {fed['hoja']}: {exc}")
            resumen.append((fed["hoja"], {'procesados': 0, 'encontrados_pnfg_exact': 0,
                                          'encontrados_pnfg_fuzzy': 0, 'encontrados_web': 0,
                                          'sin_resultado': 0, 'no_match_nombre': 0}))
        print(f"  Pausa {SLEEP_FED}s...")
        time.sleep(SLEEP_FED)

    print(f"\n{'='*70}\nRESUMEN refill\n{'='*70}")
    tot = {'procesados':0, 'exact':0, 'fuzzy':0, 'web':0, 'sin_match':0}
    for hoja, s in resumen:
        exact = s.get('encontrados_pnfg_exact', 0)
        fuzzy = s.get('encontrados_pnfg_fuzzy', 0)
        web   = s.get('encontrados_web', 0)
        proc  = s.get('procesados', 0)
        no_match = s.get('no_match_nombre', 0)
        encontrados = exact + fuzzy + web
        print(f"  {hoja:30}  total {encontrados:4}/{proc:4}  "
              f"(exact {exact}, fuzzy {fuzzy}, "
              f"web {web}, sin_match {no_match})")
        tot['procesados'] += proc
        tot['exact'] += exact
        tot['fuzzy'] += fuzzy
        tot['web']   += web
        tot['sin_match'] += no_match
    total_enc = tot['exact'] + tot['fuzzy'] + tot['web']
    pct = (total_enc/tot['procesados']*100) if tot['procesados'] else 0
    print(f"\n  TOTAL: {total_enc}/{tot['procesados']} = {pct:.1f}%")
    print(f"  Desglose: exact={tot['exact']} fuzzy={tot['fuzzy']} web={tot['web']} sin_match={tot['sin_match']}")


if __name__ == "__main__":
    main()
