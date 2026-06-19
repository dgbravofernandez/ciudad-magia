"""
Scraper unificado — Federaciones de fútbol españolas
Extrae Nombre, Email y Teléfono de cada federación y los escribe
en la hoja correspondiente de Seguimiento-Clubes-Multi-Federacion.xlsx.

Uso:
    python scraper_federaciones.py                    # todas las federaciones
    python scraper_federaciones.py "RFAF Andalucia"   # solo una
    python scraper_federaciones.py "FGF Galicia" "FAF Aragon"  # varias
"""

import os
import re
import sys
import tempfile
import time

# Forzar UTF-8 en stdout aunque esté redirigido a fichero
if hasattr(sys.stdout, 'reconfigure'):
    sys.stdout.reconfigure(encoding='utf-8', errors='replace')

import openpyxl
import requests
from bs4 import BeautifulSoup


def atomic_save(wb, path: str, retries: int = 6, delay: float = 5.0) -> None:
    """
    Guarda el workbook en un temp y renombra (evita Errno 22 en Windows).
    Reintenta si el archivo destino está bloqueado por otro proceso.
    """
    import shutil
    dir_ = os.path.dirname(os.path.abspath(path))
    fd, tmp = tempfile.mkstemp(dir=dir_, suffix='.xlsx')
    os.close(fd)
    try:
        wb.save(tmp)
    except Exception:
        try:
            os.unlink(tmp)
        except OSError:
            pass
        raise
    # Intentar reemplazar con reintentos (Windows puede tener el destino bloqueado)
    last_err = None
    for attempt in range(retries):
        try:
            os.replace(tmp, path)
            return  # éxito
        except PermissionError as e:
            last_err = e
            if attempt < retries - 1:
                time.sleep(delay * (attempt + 1))
    # Último recurso: shutil.copy2 + borrar temp
    try:
        shutil.copy2(tmp, path)
        os.unlink(tmp)
        return
    except Exception:
        pass
    try:
        os.unlink(tmp)
    except OSError:
        pass
    raise last_err  # re-raise el error original si todo falló

# ─── CONFIGURACIÓN ────────────────────────────────────────────────────────────
ruta_excel = r"C:\Users\dgbra\Ciudad Magia\docs\marketing\Seguimiento-Clubes-Multi-Federacion.xlsx"
SLEEP = 1.5        # segundos entre peticiones a fichas individuales
SLEEP_PAGE = 2.0   # segundos entre páginas del listado
SLEEP_FED  = 5.0   # segundos entre federaciones distintas

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/120.0.0.0 Safari/537.36"
    ),
    "Accept": "text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8",
    "Accept-Language": "es-ES,es;q=0.9",
}

# ─── MAPA DE FEDERACIONES ─────────────────────────────────────────────────────
#
# tipo "A" → listado con javascript:Ver(id)  + ficha NFG_VerClub
# tipo "B" → listado con href relativo NFG_VisEquipos + ficha en mismo dominio
#
FEDERACIONES = [
    # ── Tipo A ────────────────────────────────────────────────────────────────
    {
        "hoja":         "RFAF Andalucia",
        "tipo":         "A",
        "list_url":     "https://www.rfaf.es/pnfg/NPcd/NFG_Clubes",
        "ficha_url":    "https://www.rfaf.es/pnfg/NPcd/NFG_VerClub",
        "cod_primaria": "1000118",
    },
    {
        "hoja":         "FFRM Murcia",
        "tipo":         "A",
        "list_url":     "https://www.ffrm.es/pnfg/NPcd/NFG_Clubes",
        "ficha_url":    "https://www.ffrm.es/pnfg/NPcd/NFG_VerClub",
        "cod_primaria": "3001859",   # Murcia usa cod_primaria distinto
    },
    {
        "hoja":         "FGF Galicia",
        "tipo":         "A",
        "list_url":     "https://www.futgal.es/pnfg/NPcd/NFG_Clubes",
        "ficha_url":    "https://www.futgal.es/pnfg/NPcd/NFG_VerClub",
        "cod_primaria": "1000118",
    },
    {
        "hoja":         "FFCM Castilla LM",
        "tipo":         "A",
        "list_url":     "https://www.ffcm.es/pnfg/NPcd/NFG_Clubes",
        "ficha_url":    "https://www.ffcm.es/pnfg/NPcd/NFG_VerClub",
        "cod_primaria": "1000118",
    },
    {
        "hoja":         "FAF Aragon",
        "tipo":         "A",
        "list_url":     "https://www.futbolaragon.com/pnfg/NPcd/NFG_Clubes",
        "ficha_url":    "https://www.futbolaragon.com/pnfg/NPcd/NFG_VerClub",
        "cod_primaria": "1000118",
    },
    {
        "hoja":         "FFIB Baleares",
        "tipo":         "A",
        "list_url":     "https://www.ffib.es/Fed/NPcd/NFG_Clubes",   # path /Fed/
        "ficha_url":    "https://www.ffib.es/Fed/NPcd/NFG_VerClub",
        "cod_primaria": "1000108",   # Baleares usa cod_primaria distinto
    },
    {
        "hoja":         "IFCF Canarias",
        "tipo":         "A",
        "list_url":     "https://www.federacioncanariafutbol.es/pnfg/NPcd/NFG_Clubes",
        "ficha_url":    "https://www.federacioncanariafutbol.es/pnfg/NPcd/NFG_VerClub",
        "cod_primaria": "1000118",
    },
    {
        "hoja":         "FCF Cantabria",
        "tipo":         "A",
        "list_url":     "https://www.rfcf.es/pnfg/NPcd/NFG_Clubes",
        "ficha_url":    "https://www.rfcf.es/pnfg/NPcd/NFG_VerClub",
        "cod_primaria": "1000118",
    },
    {
        "hoja":         "FRF La Rioja",
        "tipo":         "A",
        "list_url":     "https://www.frfutbol.com/pnfg/NPcd/NFG_Clubes",
        "ficha_url":    "https://www.frfutbol.com/pnfg/NPcd/NFG_VerClub",
        "cod_primaria": "1000118",
    },
    {
        "hoja":         "FEXF Extremadura",
        "tipo":         "A",
        "list_url":     "https://www.fexfutbol.com/pnfg/NPcd/NFG_Clubes",
        "ficha_url":    "https://www.fexfutbol.com/pnfg/NPcd/NFG_VerClub",
        "cod_primaria": "1000118",
    },
    {
        "hoja":         "FVF Euskadi",
        "tipo":         "A",
        "list_url":     "https://webfvf.novanet.es/pnfg/NPcd/NFG_Clubes",
        "ficha_url":    "https://webfvf.novanet.es/pnfg/NPcd/NFG_VerClub",
        "cod_primaria": "1000120",
    },
    {
        "hoja":         "RFMF Melilla",
        "tipo":         "A",
        "list_url":     "https://www.rfmf.es/pnfg/NPcd/NFG_Clubes",
        "ficha_url":    "https://www.rfmf.es/pnfg/NPcd/NFG_VerClub",
        "cod_primaria": "1000118",
    },
    # ── Tipo B ────────────────────────────────────────────────────────────────
    {
        "hoja":         "FCYLF Castilla Leon",
        "tipo":         "B",
        "list_url":     "https://www.rfcylf.es/pnfg/NPcd/NFG_LstEquipos",
        "base_domain":  "https://www.rfcylf.es",
        "cod_primaria": "1000119",
    },
    {
        "hoja":         "FFPA Asturias",
        "tipo":         "B",
        "list_url":     "https://www.asturfutbol.es/pnfg/NPcd/NFG_LstEquipos",
        "base_domain":  "https://www.asturfutbol.es",
        "cod_primaria": "1000119",
    },
    {
        "hoja":         "FNF Navarra",
        "tipo":         "B",
        "list_url":     "https://www.futnavarra.es/pnfg/NPcd/NFG_LstEquipos",
        "base_domain":  "https://www.futnavarra.es",
        "cod_primaria": "1000119",
    },
]

# ─── COLUMNAS en el Excel (1-indexed) ─────────────────────────────────────────
COL_CLUB   = 1   # A
COL_EMAIL  = 4   # D
COL_TEL    = 5   # E
COL_FED    = 6   # F  Federación / Liga
COL_ESTADO = 7   # G  Estado


# ─── HELPERS ──────────────────────────────────────────────────────────────────

def make_session(base_url: str) -> requests.Session:
    """Crea una sesión con cookies inicializadas visitando la URL base."""
    s = requests.Session()
    s.headers.update(HEADERS)
    try:
        s.get(base_url, timeout=15)
    except Exception:
        pass
    return s


def get_soup(url: str, session: requests.Session | None = None) -> BeautifulSoup | None:
    try:
        getter = session if session else requests
        r = getter.get(url, headers=HEADERS, timeout=15)
        r.raise_for_status()
        r.encoding = r.apparent_encoding
        if len(r.text) < 500:   # respuesta vacía = posible bloqueo
            print(f"  [WARN] Respuesta muy corta ({len(r.text)} chars): {url}")
            return None
        return BeautifulSoup(r.text, "html.parser")
    except Exception as e:
        print(f"  [HTTP ERROR] {url}: {e}")
        return None


def parse_email(soup: BeautifulSoup) -> str:
    if not soup:
        return ""
    for h5 in soup.find_all("h5"):
        strong = h5.find("strong")
        if strong and "email" in strong.get_text(strip=True).lower():
            text = h5.get_text(separator=" ", strip=True)
            m = re.search(r"[\w.\-+]+@[\w.\-]+\.[a-z]{2,}", text, re.I)
            if m:
                return m.group(0).strip()
    return ""


def parse_phone(soup: BeautifulSoup) -> str:
    if not soup:
        return ""
    for h5 in soup.find_all("h5"):
        strong = h5.find("strong")
        if not strong:
            continue
        label = strong.get_text(strip=True).lower()
        if re.search(r"tel|tfno|telf|movil|m[oó]vil|phone|fax", label):
            text = h5.get_text(separator=" ", strip=True)
            m = re.search(r"(\+?[\d][\d\s\-\.]{6,})", text)
            if m:
                return re.sub(r"\s+", " ", m.group(1)).strip()
    return ""


def get_existing_names(ws) -> set:
    """Devuelve el set de nombres ya en la hoja (en minúsculas) para deduplicar."""
    names = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[COL_CLUB - 1]:
            names.add(str(row[COL_CLUB - 1]).strip().lower())
    return names


COL_UBICACION = 2   # B

def append_club(ws, nombre: str, email: str, phone: str, fed_nombre: str, ubicacion: str = "") -> None:
    """Escribe una fila al final de la hoja. NO guarda — llamar atomic_save después de cada página."""
    next_row = ws.max_row + 1
    ws.cell(row=next_row, column=COL_CLUB).value      = nombre
    ws.cell(row=next_row, column=COL_UBICACION).value = ubicacion or None
    ws.cell(row=next_row, column=COL_EMAIL).value     = email or None
    ws.cell(row=next_row, column=COL_TEL).value       = phone or None
    ws.cell(row=next_row, column=COL_FED).value       = fed_nombre
    ws.cell(row=next_row, column=COL_ESTADO).value    = "Nuevo"


# ─── ESTRATEGIA A: javascript:Ver(id) ─────────────────────────────────────────

def scrape_tipo_a(fed: dict, ws, wb) -> int:
    existing = get_existing_names(ws)
    total = 0
    page = 1
    session = make_session(fed["list_url"] + f"?cod_primaria={fed['cod_primaria']}")
    # Detección de rate-limiting: si N fichas consecutivas devuelven vacío, modo solo-nombre
    consecutive_empty = 0
    EMPTY_THRESHOLD = 8
    skip_fichas = False

    while True:
        url = (
            f"{fed['list_url']}"
            f"?cod_primaria={fed['cod_primaria']}"
            f"&NPcd_PageLines=999"
            f"&NPcd_Page={page}"
            f"&Buscar=1"
        )
        soup = get_soup(url, session)
        if not soup:
            break

        found_on_page = 0   # todos los clubs encontrados en la página (new + skip)
        for a in soup.find_all("a", href=re.compile(r"javascript:Ver\(\d+\)", re.I)):
            nombre = a.get_text(strip=True)
            if not nombre:
                continue

            found_on_page += 1

            if nombre.lower() in existing:
                continue   # ya en Excel, no imprimir SKIP para no saturar el log

            m = re.search(r"Ver\((\d+)\)", a["href"], re.I)
            if not m:
                continue
            codigo = m.group(1)

            # Extraer localidad y provincia de la fila del listado (sin coste extra)
            localidad = ""
            tr = a.find_parent("tr")
            if tr:
                tds = tr.find_all("td")
                # Estructura típica: vacío | código | nombre | localidad | provincia | equipos
                if len(tds) >= 5:
                    localidad_raw = tds[3].get_text(strip=True)
                    provincia_raw = tds[4].get_text(strip=True)
                    if localidad_raw and provincia_raw and localidad_raw != provincia_raw:
                        localidad = f"{localidad_raw} ({provincia_raw})"
                    elif localidad_raw:
                        localidad = localidad_raw

            email, phone = "", ""
            if not skip_fichas:
                time.sleep(SLEEP)
                ficha_url = (
                    f"{fed['ficha_url']}"
                    f"?cod_primaria={fed['cod_primaria']}"
                    f"&codigo_club={codigo}"
                )
                soup_ficha = get_soup(ficha_url, session)
                if soup_ficha is None:
                    consecutive_empty += 1
                    if consecutive_empty >= EMPTY_THRESHOLD:
                        print(f"  [RATE-LIMIT] {consecutive_empty} fichas vacias - modo solo-nombre activado")
                        skip_fichas = True
                else:
                    consecutive_empty = 0
                    email = parse_email(soup_ficha)
                    phone = parse_phone(soup_ficha)

            append_club(ws, nombre, email, phone, fed["hoja"], localidad)
            existing.add(nombre.lower())
            total += 1
            print(f"  OK {nombre:<50} email: {email or '-':<35} tel: {phone or '-'}")

        # Parar solo cuando la página no tiene ningún club (fin real de paginación)
        if found_on_page == 0:
            break
        # Guardar tras cada página (no tras cada club) — evita Errno 22
        atomic_save(wb, ruta_excel)
        page += 1
        time.sleep(SLEEP_PAGE)

    return total


# ─── ESTRATEGIA B: href relativo NFG_VisEquipos ───────────────────────────────

def scrape_tipo_b(fed: dict, ws, wb) -> int:
    existing = get_existing_names(ws)
    total = 0
    page = 1
    session = make_session(fed["list_url"] + f"?cod_primaria={fed['cod_primaria']}")

    while True:
        url = (
            f"{fed['list_url']}"
            f"?cod_primaria={fed['cod_primaria']}"
            f"&NPcd_PageLines=999"
            f"&NPcd_Page={page}"
            f"&Buscar=1"
        )
        soup = get_soup(url, session)
        if not soup:
            break

        found_on_page = 0
        for a in soup.find_all("a", href=re.compile(r"NFG_VisEquipos", re.I)):
            nombre = a.get_text(strip=True)
            if not nombre:
                continue

            found_on_page += 1

            if nombre.lower() in existing:
                continue

            href = a.get("href", "")
            if not href.startswith("http"):
                href = fed["base_domain"] + href

            time.sleep(SLEEP)
            soup_ficha = get_soup(href, session)
            email = parse_email(soup_ficha)
            phone = parse_phone(soup_ficha)

            append_club(ws, nombre, email, phone, fed["hoja"])
            existing.add(nombre.lower())
            total += 1
            print(f"  OK {nombre:<50} email: {email or '-':<35} tel: {phone or '-'}")

        if found_on_page == 0:
            break
        # Guardar tras cada página (no tras cada club) — evita Errno 22
        atomic_save(wb, ruta_excel)
        page += 1
        time.sleep(SLEEP_PAGE)

    return total


# ─── MAIN ─────────────────────────────────────────────────────────────────────

def main():
    if ruta_excel == "RUTA_DEL_ARCHIVO":
        print("[WARN] Edita la variable ruta_excel con la ruta real al .xlsx")
        return

    # Filtro por argumento de línea de comandos
    filtro = [a.lower() for a in sys.argv[1:]]

    wb = openpyxl.load_workbook(ruta_excel)
    resumen = []

    for fed in FEDERACIONES:
        hoja = fed["hoja"]

        if filtro and not any(f in hoja.lower() for f in filtro):
            continue

        if hoja not in wb.sheetnames:
            print(f"\n[WARN] Hoja '{hoja}' no encontrada en el Excel - saltando.")
            continue

        print(f"\n{'='*70}")
        print(f">> {hoja}  (tipo {fed['tipo']})")
        print(f"{'='*70}")

        ws = wb[hoja]

        try:
            if fed["tipo"] == "A":
                n = scrape_tipo_a(fed, ws, wb)
            else:
                n = scrape_tipo_b(fed, ws, wb)
            atomic_save(wb, ruta_excel)  # guardado final de la federación
            resumen.append((hoja, n, "OK"))
        except Exception as e:
            print(f"  [FATAL] {hoja}: {e}")
            resumen.append((hoja, 0, f"ERROR: {e}"))

        print(f"  Pausa {SLEEP_FED}s antes de la siguiente federacion..."  )
        time.sleep(SLEEP_FED)

    print(f"\n{'='*70}")
    print("RESUMEN")
    print(f"{'='*70}")
    for hoja, n, estado in resumen:
        print(f"  {estado}  {hoja:<30}  {n} clubes añadidos")


if __name__ == "__main__":
    main()
